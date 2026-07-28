#!/usr/bin/env python3
"""
Refresh the "Estimated Value" column of a card XLSX from live TCGdex prices.

This is an out-of-band step: run it on a copy of an event's XLSX BEFORE the
normal ingest flow. It does not touch Algolia — it only rewrites values in the
spreadsheet, so the standard setup_event.sh / ingest.py pipeline stays unchanged.

For each in-machine card it resolves the TCGdex set + card by reusing ingest.py's
set-resolution and card-number normalization, reads
pricing.tcgplayer.<variant>.marketPrice (USD), and writes the refreshed value back.
Unlike ingest.py, it does NOT auto-resolve a missing card Number from the Pokemon
Name — rows without a usable Number are skipped and logged. When a card was
printed in multiple finishes
(normal / reverse-holofoil / holofoil), it picks the variant whose *current*
price is closest to the sheet's *existing* value — using the original valuation
as the finish selector, then refreshing the number to today's market.

Cards that can't be resolved (set/card not found, no TCGPlayer pricing) keep
their existing value and are logged for review.

Usage:
    python refresh_values.py <xlsx_path> [--dry-run]

Examples:
    python refresh_values.py "../data-files/etail-boston-2026/TCG Search Raw List - eTail Boston 2026.xlsx" --dry-run
    python refresh_values.py "../data-files/etail-boston-2026/TCG Search Raw List - eTail Boston 2026.xlsx"
"""

import re
import sys
import time
import argparse
from pathlib import Path

import openpyxl
import requests

TCGDEX_BASE_URL = "https://api.tcgdex.net/v2/en"
REQUIRED_COLUMNS = {"Pokemon Name", "Number", "# in Machine", "Card Type", "Estimated Value"}

# Cell values ingest.py treats as missing — blanks plus the 'nan'/'---' sentinels (ingest.py:91).
MISSING_CELL_VALUES = {"", "nan", "---"}


# ── Parsing helpers (mirror ingest.py semantics) ────────────────────────────────

def is_missing(raw) -> bool:
    """True for blank cells and the sentinels ingest.py treats as missing ('nan', '---')."""
    if raw is None:
        return True
    if isinstance(raw, float) and raw != raw:  # float NaN
        return True
    return str(raw).strip().lower() in MISSING_CELL_VALUES

def extract_card_set_from_sheet_name(sheet_name: str) -> str:
    """Extract set name from a sheet tab (same rules as ingest.py)."""
    name = re.sub(r"\s*\(\d+\)?\s*$", "", sheet_name).strip()
    parts = re.split(r"\s+-\s*", name)
    card_set = parts[-1].strip().replace("_", ": ")
    return " ".join(card_set.split())


def normalize_number(raw) -> str | None:
    """Normalize a Number cell to a bare local id, e.g. 287.0 / '287/217' -> '287'. Missing -> None."""
    if is_missing(raw):
        return None
    s = str(int(raw)) if isinstance(raw, float) else str(raw).strip()
    return s.split("/")[0].strip()


def parse_value(raw) -> float | None:
    """Parse an existing Estimated Value — native float (XLSX) or '$20.60' (CSV). Missing -> None."""
    if is_missing(raw):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def variant_prices(tcgplayer: dict) -> dict:
    """Map {variant: price} from a pricing.tcgplayer block (marketPrice, else midPrice)."""
    prices = {}
    for key, val in (tcgplayer or {}).items():
        if key in ("unit", "updated") or not isinstance(val, dict):
            continue
        price = val.get("marketPrice")
        if price is None:
            price = val.get("midPrice")
        if price is not None:
            prices[key] = float(price)
    return prices


def choose_price(prices: dict, existing: float | None) -> tuple[str | None, float | None]:
    """Pick the variant nearest the existing value; fall back to the priciest."""
    if not prices:
        return None, None
    if existing is not None:
        variant = min(prices, key=lambda k: abs(prices[k] - existing))
    else:
        variant = max(prices, key=lambda k: prices[k])
    return variant, prices[variant]


# ── TCGdex lookups ──────────────────────────────────────────────────────────────

class Tcgdex:
    """Small TCGdex client with set-list + set-id caching."""

    def __init__(self):
        self.session = requests.Session()
        self._sets = None            # cached /sets list
        self._set_id: dict[str, str | None] = {}  # set_name -> id

    def _get(self, path: str) -> dict | list | None:
        for attempt in range(3):
            try:
                r = self.session.get(f"{TCGDEX_BASE_URL}{path}", timeout=10)
                r.raise_for_status()
                return r.json()
            except requests.exceptions.RequestException as e:
                if attempt < 2:
                    time.sleep(1 * (attempt + 1))
                else:
                    print(f"  ✗ TCGdex request failed for {path}: {e}")
                    return None

    def _all_sets(self) -> list:
        if self._sets is None:
            self._sets = self._get("/sets") or []
        return self._sets

    def resolve_set_id(self, set_name: str) -> str | None:
        """Resolve a sheet set name to a TCGdex set id (same cascade as ingest.py)."""
        if set_name in self._set_id:
            return self._set_id[set_name]

        def norm(n: str) -> str:
            return re.sub(r"[^a-z0-9\s]", "", n.lower())

        sets = self._all_sets()
        candidates = [set_name]
        if ":" in set_name:
            candidates.append(set_name.split(":", 1)[1].strip())

        set_id = None
        # exact / without-prefix match
        for cand in candidates:
            for s in sets:
                if s.get("name", "").lower() == cand.lower():
                    set_id = s.get("id")
                    break
            if set_id:
                break
        # normalized substring match
        if not set_id:
            target = norm(set_name)
            for s in sets:
                if target in norm(s.get("name", "")):
                    set_id = s.get("id")
                    break
        # suffix fallback for truncated (31-char) tab names
        if not set_id:
            words = set_name.split()
            for i in range(1, len(words)):
                suffix = norm(" ".join(words[i:]))
                for s in sets:
                    if suffix in norm(s.get("name", "")):
                        set_id = s.get("id")
                        break
                if set_id:
                    break

        self._set_id[set_name] = set_id
        return set_id

    def card_tcgplayer_prices(self, set_id: str, number: str) -> dict:
        """Return {variant: price} for a card, or {} if unavailable."""
        card = self._get(f"/sets/{set_id}/{number}")
        if not isinstance(card, dict):
            return {}
        return variant_prices((card.get("pricing") or {}).get("tcgplayer") or {})


# ── Sheet processing ────────────────────────────────────────────────────────────

def header_map(ws) -> dict:
    """Map stripped header name -> 1-based column index from row 1."""
    cols = {}
    for cell in ws[1]:
        if cell.value is not None:
            cols[str(cell.value).strip()] = cell.column
    return cols


def refresh_workbook(path: Path, dry_run: bool, max_ratio: float) -> int:
    wb = openpyxl.load_workbook(path)
    tcg = Tcgdex()

    updated = unchanged = skipped_rows = 0
    failures: list[str] = []
    suspicious: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = header_map(ws)
        if not REQUIRED_COLUMNS.issubset(cols.keys()):
            continue  # not a card set (Landing Page Sections, Most expensive, ...)

        set_name = extract_card_set_from_sheet_name(sheet_name)
        set_id = tcg.resolve_set_id(set_name)
        print(f"\n=== {sheet_name}  ->  set '{set_name}'  (tcgdex: {set_id or 'UNRESOLVED'}) ===")

        name_c = cols["Pokemon Name"]
        num_c = cols["Number"]
        qty_c = cols["# in Machine"]
        val_c = cols["Estimated Value"]

        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=name_c).value
            if is_missing(name):
                continue
            name = str(name).strip()

            qty = ws.cell(row=r, column=qty_c).value
            if is_missing(qty):
                continue  # not in machine (blank/nan/---) — ingest skips it, so do we

            number = normalize_number(ws.cell(row=r, column=num_c).value)
            existing = parse_value(ws.cell(row=r, column=val_c).value)

            if not set_id or not number:
                skipped_rows += 1
                failures.append(f"{sheet_name}: {name} #{number} — {'set unresolved' if not set_id else 'missing number'}")
                continue

            prices = tcg.card_tcgplayer_prices(set_id, number)
            variant, new_val = choose_price(prices, existing)
            if new_val is None:
                skipped_rows += 1
                failures.append(f"{sheet_name}: {name} #{number} — no TCGPlayer pricing (kept {existing})")
                continue

            new_val = round(new_val, 2)
            old_str = f"{existing:.2f}" if existing is not None else "—"

            # Guard against number collisions / missing premium variants: if the
            # nearest price is wildly off the existing value, the sheet's number
            # likely didn't resolve to the same physical card. Keep the old value
            # and flag it rather than silently writing a bogus price.
            if existing and existing > 0 and not (existing / max_ratio <= new_val <= existing * max_ratio):
                suspicious.append(f"{sheet_name}: {name} #{number} — kept {old_str}, TCGdex nearest was {new_val:.2f} [{variant}]")
                print(f"  ⚠ {name} #{number}: {old_str} -> {new_val:.2f} [{variant}]  SUSPICIOUS — kept {old_str}")
                continue

            if existing is not None and new_val == round(existing, 2):
                unchanged += 1
                print(f"  = {name} #{number}: {old_str} (unchanged) [{variant}]")
                continue

            updated += 1
            print(f"  → {name} #{number}: {old_str} -> {new_val:.2f} [{variant}]")
            if not dry_run:
                ws.cell(row=r, column=val_c).value = new_val

    print("\n" + "=" * 60)
    print(f"Updated: {updated}   Unchanged: {unchanged}   "
          f"Suspicious (kept): {len(suspicious)}   No pricing (kept): {skipped_rows}")
    if suspicious:
        print(f"\nSuspicious — refreshed price deviated >{max_ratio}x from existing "
              f"(likely a number mismatch; kept existing value):")
        for s in suspicious:
            print(f"  ⚠ {s}")
    if failures:
        print("\nNo TCGdex/TCGPlayer pricing — kept existing value:")
        for f in failures:
            print(f"  ⚠ {f}")

    if dry_run:
        print("\n(--dry-run: no changes written)")
    else:
        wb.save(path)
        print(f"\n✓ Saved refreshed values to {path.name}")

    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Refresh Estimated Value from TCGdex prices")
    parser.add_argument("xlsx", help="Path to the card XLSX to refresh (edited in place)")
    parser.add_argument("--dry-run", action="store_true", help="Print old->new diffs without writing")
    parser.add_argument("--max-ratio", type=float, default=4.0,
                        help="Flag (and keep existing) when the refreshed price deviates by more "
                             "than this factor from the existing value (default: 4.0)")
    args = parser.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        print(f"ERROR: file not found: {path}")
        return 1

    return refresh_workbook(path, args.dry_run, args.max_ratio)


if __name__ == "__main__":
    sys.exit(main())
